import os
import sys
import json
import io
import time
import datetime
import urllib.parse
import threading
import shutil
from http.server import ThreadingHTTPServer, SimpleHTTPRequestHandler

# Import Windows COM & Win32 File API
try:
    import win32com.client
    import win32file
    import win32con
    import pythoncom
    HAS_WIN32 = True
except ImportError:
    HAS_WIN32 = False

import openpyxl
import hashlib

# Set output encoding to UTF-8
try:
    sys.stdout.reconfigure(encoding='utf-8', line_buffering=True)
except Exception:
    pass

if getattr(sys, 'frozen', False):
    APP_DIR = os.path.dirname(sys.executable)
else:
    APP_DIR = os.path.dirname(os.path.abspath(__file__))

CONFIG_FILE = os.path.join(APP_DIR, "config.json")
PORT = 8080

DEFAULT_CONFIG = {
    "excel_path": "",
    "excel_paths": {
        "平日": "",
        "月曜": "",
        "火曜": "",
        "日・祝": ""
    },
    "poll_interval_sec": 5,
    "font_size_scale": 1.0,
    "auto_scroll_speed": 40,
    "scroll_speed_px_per_sec": 50,
    "bottom_pause_sec": 4,
    "top_pause_sec": 2,
    "theme": "dark",
    "export_dir": ""
}

DAYS_ORDER = ["平日", "月曜", "火曜", "日・祝"]

DAY_ALIASES = {
    "平日": "平日", "heijitsu": "平日", "weekday": "平日", "水": "平日", "木": "平日", "金": "平日", "土": "平日", "wed": "平日", "thu": "平日", "fri": "平日", "sat": "平日",
    "月曜": "月曜", "月": "月曜", "mon": "月曜", "monday": "月曜",
    "火曜": "火曜", "火": "火曜", "tue": "火曜", "tuesday": "火曜",
    "日・祝": "日・祝", "日祝": "日・祝", "日": "日・祝", "祝": "日・祝", "sun": "日・祝", "sunday": "日・祝", "holiday": "日・祝"
}

DAY_KEYWORD_MAP = {
    "平日": ["(平日)"],
    "月曜": ["(月)", "(月曜)"],
    "火曜": ["(火)", "(火曜)"],
    "日・祝": ["(日祝)", "(日・祝)"]
}

# In-memory thread-safe cache
MEMORY_CACHE = {}
LAST_FILE_MTIME = {}
CACHE_LOCK = threading.Lock()


def resolve_canonical_day(day_input):
    try:
        if not day_input or not str(day_input).strip():
            weekday = datetime.datetime.now().weekday()
            # Monday=0, Tuesday=1, Wednesday=2, Thursday=3, Friday=4, Saturday=5, Sunday=6
            if weekday == 6:
                return "日・祝"
            elif weekday == 0:
                return "月曜"
            elif weekday == 1:
                return "火曜"
            else:
                return "平日"

        clean = str(day_input).strip().lower()
        if clean in DAY_ALIASES:
            return DAY_ALIASES[clean]
        for k, v in DAY_ALIASES.items():
            if k.lower() == clean:
                return v
        for k, v in DAY_ALIASES.items():
            if k.lower() in clean:
                return v
        return "平日"
    except Exception:
        return "平日"


def load_config():
    try:
        if not os.path.exists(CONFIG_FILE):
            cfg = dict(DEFAULT_CONFIG)
            save_config(cfg)
            return cfg
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            cfg = json.load(f)
            merged = {**DEFAULT_CONFIG, **cfg}
            if "excel_paths" not in merged or not isinstance(merged["excel_paths"], dict):
                merged["excel_paths"] = dict(DEFAULT_CONFIG["excel_paths"])
            return merged
    except Exception:
        return dict(DEFAULT_CONFIG)


def save_config(cfg):
    try:
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(cfg, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print("[CONFIG SAVE ERROR]", e, flush=True)


def expand_path(p):
    """Expands %USERPROFILE%, %USERNAME%, and environment variables in path."""
    if not p or not isinstance(p, str):
        return ""
    expanded = os.path.expandvars(p.strip())
    return os.path.normpath(expanded)


def get_candidate_onedrive_roots():
    candidate_onedrive_roots = []
    user_profile = os.environ.get("USERPROFILE", "")
    onedrive_env = os.environ.get("OneDriveCommercial") or os.environ.get("OneDrive")
    if onedrive_env and os.path.exists(onedrive_env):
        candidate_onedrive_roots.append(onedrive_env)
    if user_profile:
        p1 = os.path.join(user_profile, "OneDrive - トヨタモビリティパーツ株式会社")
        if os.path.exists(p1) and p1 not in candidate_onedrive_roots:
            candidate_onedrive_roots.append(p1)
    for fallback in [
        r"C:\Users\85371-housen-k5\OneDrive - トヨタモビリティパーツ株式会社",
        r"c:\Users\00137184\OneDrive - トヨタモビリティパーツ株式会社"
    ]:
        if os.path.exists(fallback) and fallback not in candidate_onedrive_roots:
            candidate_onedrive_roots.append(fallback)
    return candidate_onedrive_roots


# Cache to log source transitions only when changed
LAST_LOGGED_SOURCE = {}


def find_excel_file_for_day(canonical_day):
    """
    Finds target Excel file with multi-tier automatic fallback:
      Priority 1: Dedicated Server PC (85371-housen-k5) direct path
      Priority 2: config.json defined candidate paths (supports list & %USERPROFILE%)
      Priority 3: Auto-detected active OneDrive / Shortcuts / ★入力シート on current running PC
    """
    global LAST_LOGGED_SOURCE
    try:
        cfg = load_config()
        current_user = os.environ.get("USERNAME", "ローカルユーザー")
        keywords = DAY_KEYWORD_MAP.get(canonical_day, ["(平日)"])

        # -------------------------------------------------------------
        # Priority 1: Check Dedicated Server PC (85371-housen-k5) paths
        # -------------------------------------------------------------
        server_pc_candidates = {
            "平日": [
                r"C:\Users\85371-housen-k5\OneDrive - トヨタモビリティパーツ株式会社\新体制移行の情報共有 - 平日(水～土)（本番用・使用不可）\(平日)作業進捗管理データ.xlsm",
                r"C:\Users\85371-housen-k5\OneDrive - トヨタモビリティパーツ株式会社\新体制移行の情報共有 - ★入力シート\平日(水～土)（本番用）\(平日)作業進捗管理データ.xlsm",
                r"C:\Users\85371-housen-k5\OneDrive - トヨタモビリティパーツ株式会社\Shortcuts\新体制移行の情報共有 - ★入力シート\平日(水～土)（本番用）\(平日)作業進捗管理データ.xlsm"
            ],
            "月曜": [
                r"C:\Users\85371-housen-k5\OneDrive - トヨタモビリティパーツ株式会社\Shortcuts\新体制移行の情報共有 - 月曜（本番用・使用不可）\(月)作業進捗管理データ.xlsm",
                r"C:\Users\85371-housen-k5\OneDrive - トヨタモビリティパーツ株式会社\Shortcuts\新体制移行の情報共有 - ★入力シート\月曜（本番用）\(月)作業進捗管理データ.xlsm"
            ],
            "火曜": [
                r"C:\Users\85371-housen-k5\OneDrive - トヨタモビリティパーツ株式会社\Shortcuts\新体制移行の情報共有 - 火曜（本番用・使用不可）\(火)作業進捗管理データ.xlsm",
                r"C:\Users\85371-housen-k5\OneDrive - トヨタモビリティパーツ株式会社\Shortcuts\新体制移行の情報共有 - ★入力シート\火曜（本番用）\(火)作業進捗管理データ.xlsm"
            ],
            "日・祝": [
                r"C:\Users\85371-housen-k5\OneDrive - トヨタモビリティパーツ株式会社\Shortcuts\新体制移行の情報共有 - 日祝（本番用・使用不可）\(日祝)作業進捗管理データ.xlsm",
                r"C:\Users\85371-housen-k5\OneDrive - トヨタモビリティパーツ株式会社\Shortcuts\新体制移行の情報共有 - ★入力シート\日祝（本番用）\(日祝)作業進捗管理データ.xlsm"
            ]
        }

        for candidate in server_pc_candidates.get(canonical_day, []):
            p = expand_path(candidate)
            if p and os.path.exists(p):
                if LAST_LOGGED_SOURCE.get(canonical_day) != p:
                    LAST_LOGGED_SOURCE[canonical_day] = p
                    print(f"[DATA SOURCE: 優先①] [{canonical_day}] サーバーPC(85371-housen-k5)のOneDriveを検出: {p}", flush=True)
                return p

        # -------------------------------------------------------------
        # Priority 2: Configured paths in config.json
        # -------------------------------------------------------------
        configured_entry = cfg.get("excel_paths", {}).get(canonical_day)
        candidates_from_config = []
        if isinstance(configured_entry, list):
            candidates_from_config.extend(configured_entry)
        elif isinstance(configured_entry, str) and configured_entry.strip():
            candidates_from_config.append(configured_entry.strip())

        if canonical_day == "平日":
            legacy_p = cfg.get("excel_path", "")
            if legacy_p and legacy_p not in candidates_from_config:
                candidates_from_config.append(legacy_p)

        for raw_p in candidates_from_config:
            p = expand_path(raw_p)
            if p and os.path.exists(p):
                if LAST_LOGGED_SOURCE.get(canonical_day) != p:
                    LAST_LOGGED_SOURCE[canonical_day] = p
                    print(f"[DATA SOURCE: 優先②] [{canonical_day}] config.json指定パスを使用: {p}", flush=True)
                return p

        # -------------------------------------------------------------
        # Priority 3: Automatic fallback on current PC's OneDrive / Shortcuts
        # -------------------------------------------------------------
        search_dirs = []
        candidate_onedrive_roots = get_candidate_onedrive_roots()

        for onedrive_root in candidate_onedrive_roots:
            if not os.path.exists(onedrive_root):
                continue
            try:
                for root_dir, subdirs, _ in os.walk(onedrive_root):
                    rel = os.path.relpath(root_dir, onedrive_root)
                    depth = len(rel.split(os.sep)) if rel != "." else 0
                    if depth <= 3:
                        if root_dir not in search_dirs:
                            if "入力シート" in root_dir or "新体制" in root_dir:
                                search_dirs.insert(0, root_dir)
                            else:
                                search_dirs.append(root_dir)
                    else:
                        subdirs.clear()
            except Exception:
                pass

        matched_candidates = []

        for s_dir in search_dirs:
            if not s_dir or not os.path.exists(s_dir):
                continue
            try:
                for fname in os.listdir(s_dir):
                    if fname.endswith((".xlsx", ".xlsm")) and not fname.startswith("~$"):
                        if any(ng in fname for ng in ["トライアル", "テスト", "パワポ表示用", "コピー"]):
                            continue
                        if canonical_day == "月曜" and "(平日)" in fname:
                            continue
                        if any(kw in fname for kw in keywords):
                            full_file = os.path.join(s_dir, fname)
                            mtime = os.path.getmtime(full_file)
                            is_official = ("作業進捗管理データ" in fname)
                            matched_candidates.append((is_official, mtime, full_file))
            except Exception:
                pass

        if matched_candidates:
            matched_candidates.sort(key=lambda x: (x[0], x[1]), reverse=True)
            chosen = matched_candidates[0][2]
            if LAST_LOGGED_SOURCE.get(canonical_day) != chosen:
                LAST_LOGGED_SOURCE[canonical_day] = chosen
                print(f"[PATH FALLBACK: 優先③] [{canonical_day}] サーバーPC未検出のため、{current_user}のOneDriveを自動検知して使用: {chosen}", flush=True)
            return chosen

        return ""
    except Exception as e:
        print(f"[PATH ERROR] [{canonical_day}]: {e}", flush=True)
        return ""


def read_locked_file_bytes(filepath):
    """Reads bytes from a file even if locked by Excel for writing."""
    if HAS_WIN32:
        try:
            handle = win32file.CreateFile(
                filepath,
                win32con.GENERIC_READ,
                win32con.FILE_SHARE_READ | win32con.FILE_SHARE_WRITE | win32con.FILE_SHARE_DELETE,
                None,
                win32con.OPEN_EXISTING,
                win32con.FILE_ATTRIBUTE_NORMAL,
                None
            )
            file_size = win32file.GetFileSize(handle)
            _, data = win32file.ReadFile(handle, file_size)
            win32file.CloseHandle(handle)
            return data
        except Exception:
            pass
    try:
        with open(filepath, "rb") as f:
            return f.read()
    except Exception:
        return None


def get_live_com_rows(target_excel_path, canonical_day):
    """Reads in-memory unsaved data from running Excel process via Windows COM strictly matching filename or day keywords."""
    if not HAS_WIN32:
        return None
    try:
        pythoncom.CoInitialize()
        wb = None
        target_name = os.path.basename(target_excel_path).lower()
        keywords = DAY_KEYWORD_MAP.get(canonical_day, [])

        try:
            xl_app = win32com.client.GetActiveObject("Excel.Application")
            if xl_app:
                # 1. Exact match on FullName or Name
                for w in xl_app.Workbooks:
                    try:
                        if w.FullName.lower() == target_excel_path.lower() or w.Name.lower() == target_name:
                            wb = w
                            break
                    except Exception:
                        pass
                # 2. Keyword match (e.g. "(日祝)" in w.Name)
                if wb is None:
                    for w in xl_app.Workbooks:
                        try:
                            w_name = w.Name.lower()
                            if any(kw.lower() in w_name for kw in keywords):
                                wb = w
                                break
                        except Exception:
                            pass
        except Exception:
            pass

        if wb is None:
            return None

        # Worksheet 1: 表示
        ws_disp = wb.Worksheets(1)
        # Worksheet 2: データ
        ws_data = None
        for s in wb.Worksheets:
            if "データ" in s.Name:
                ws_data = s
                break
        if ws_data is None:
            ws_data = ws_disp

        disp_range = ws_disp.Range("A1:CZ145").Value
        data_range = ws_data.Range("A1:CZ145").Value

        disp_rows = [list(r) for r in disp_range]
        data_rows = [list(r) for r in data_range]

        return {
            "title": ws_disp.Name,
            "disp_rows": disp_rows,
            "data_rows": data_rows,
            "last_modified": datetime.datetime.now().strftime("%Y/%m/%d %H:%M:%S") + " (Live RAM)"
        }
    except Exception:
        return None
    finally:
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass


def format_cell_name(val):
    if val is None:
        return ""
    if isinstance(val, float):
        if val.is_integer():
            return str(int(val))
        return str(val)
    if isinstance(val, int):
        return str(val)
    s = str(val).strip()
    if s.endswith(".0"):
        try:
            float(s)
            s = s[:-2]
        except ValueError:
            pass
    return s


def format_cell_time(time_val):
    if time_val is None:
        return ""
    if isinstance(time_val, (datetime.time, datetime.datetime)):
        return time_val.strftime("%H:%M")
    if isinstance(time_val, (int, float)):
        total_minutes = int(round(time_val * 24 * 60))
        hours = (total_minutes // 60) % 24
        minutes = total_minutes % 60
        return f"{hours:02d}:{minutes:02d}"
    s = str(time_val).strip()
    if " " in s:
        s = s.split(" ")[-1]
    if len(s) >= 4 and ":" in s:
        parts = s.split(":")
        try:
            return f"{int(parts[0]):02d}:{int(parts[1]):02d}"
        except ValueError:
            return s
    return s


def parse_rows_into_signage_data(canonical_day, excel_path, disp_rows, data_rows, last_mod):
    courses = []
    last_vehicle = ""

    for r_idx in range(2, len(disp_rows), 2):
        r1 = disp_rows[r_idx]
        r2 = disp_rows[r_idx + 1] if r_idx + 1 < len(disp_rows) else [None] * 30

        v_val = r1[0] if len(r1) > 0 and r1[0] is not None else None
        c_val = r1[1] if len(r1) > 1 and r1[1] is not None else None
        time_val = r1[2] if len(r1) > 2 and r1[2] is not None else None

        if v_val is None and c_val is None and time_val is None:
            continue

        v_str = format_cell_name(v_val)
        if v_str != "":
            last_vehicle = v_str

        course_name = format_cell_name(c_val)
        time_str = format_cell_time(time_val)

        data_r1 = data_rows[r_idx] if r_idx < len(data_rows) else [None] * 105
        data_r2 = data_rows[r_idx + 1] if r_idx + 1 < len(data_rows) else [None] * 105

        # 1. 振出: 奇数行 (data_r1) の G..AE列 (インデックス 6..30)
        furidashi_items = []
        for num_idx in range(1, 26):
            data_col_idx = 5 + num_idx
            st_val = data_r1[data_col_idx] if len(data_r1) > data_col_idx and data_r1[data_col_idx] is not None else 0
            val_num = int(st_val) if isinstance(st_val, (int, float)) else 0
            status = val_num if val_num in (99, 1) else 0
            furidashi_items.append({
                "num": num_idx,
                "status": status
            })

        # 2. 査照: 偶数行 (data_r2) の G..AE列 (インデックス 6..30)
        sagyo_items = []
        for num_idx in range(1, 26):
            data_col_idx = 5 + num_idx
            st_val = data_r2[data_col_idx] if len(data_r2) > data_col_idx and data_r2[data_col_idx] is not None else 0
            val_num = int(st_val) if isinstance(st_val, (int, float)) else 0
            status = val_num if val_num in (99, 1) else 0
            sagyo_items.append({
                "num": num_idx,
                "status": status
            })

        # 3. 伝票: データシート CH列 (インデックス 85) が 1
        slip_val = data_r1[85] if len(data_r1) > 85 and data_r1[85] is not None else 0
        is_slip_done = (slip_val in (1, "1", 1.0) or slip_val is True)

        # 4. 配送なし (No Delivery) の判定
        # 条件: データシートの (G～AE列が0 かつ AF列が99) または 表示シートの (AF列がTrue)
        all_furidashi_zero = all(item["status"] == 0 for item in furidashi_items)
        data_af_val = data_r1[31] if len(data_r1) > 31 else None
        disp_af_val = r1[31] if len(r1) > 31 else None

        is_data_af_99 = (data_af_val in (99, "99", 99.0) or str(data_af_val).strip() == "99")
        is_disp_af_true = (disp_af_val is True or str(disp_af_val).strip().upper() in ("TRUE", "1"))

        is_no_delivery = (all_furidashi_zero and is_data_af_99) or is_disp_af_true

        # 5. コース完了判定 (データシート CK列=1 または 配送なし)
        is_course_completed = False
        if len(data_r1) > 88 and data_r1[88] is not None:
            ck_val = str(data_r1[88]).strip()
            if ck_val in ("1", "1.0") or data_r1[88] == 1 or data_r1[88] is True:
                is_course_completed = True
        if is_no_delivery:
            is_course_completed = True

        # 6. 集約コース完了時間 (CM列: インデックス 90) & 時間差 (CN列: インデックス 91)
        cm_val = data_r1[90] if len(data_r1) > 90 else None
        cn_val = data_r1[91] if len(data_r1) > 91 else None
        group_completed_time = format_cell_time(cm_val)

        group_diff_min = None
        if cn_val is not None and str(cn_val).strip() != "":
            try:
                group_diff_min = int(round(float(cn_val)))
            except (ValueError, TypeError):
                group_diff_min = None

        course_id = f"course_{r_idx//2 + 1}"

        courses.append({
            "id": course_id,
            "row_index": r_idx + 1,
            "vehicle": last_vehicle,
            "course": course_name,
            "time": time_str,
            "is_completed": is_course_completed,
            "is_no_delivery": is_no_delivery,
            "group_completed_time": group_completed_time,
            "group_diff_minutes": group_diff_min,
            "furidashi": {
                "label": "振出",
                "items": furidashi_items
            },
            "sagyo": {
                "label": "査照",
                "items": sagyo_items
            },
            "slip": {
                "label": "伝票",
                "is_done": is_slip_done
            }
        })

    return {
        "success": True,
        "day": canonical_day,
        "title": canonical_day,
        "excel_file": os.path.basename(excel_path),
        "count": len(courses),
        "last_modified": last_mod,
        "courses": courses
    }


def refresh_data_for_day(canonical_day):
    """Background update for a single day. Returns data object."""
    try:
        excel_path = find_excel_file_for_day(canonical_day)
        if not excel_path or not os.path.exists(excel_path):
            with CACHE_LOCK:
                cached = MEMORY_CACHE.get(canonical_day)
                if cached:
                    return cached
            return {
                "success": False,
                "day": canonical_day,
                "error": f"「{canonical_day}」のExcelファイルが見つかりません"
            }

        # 1. Try Live RAM COM
        com_result = get_live_com_rows(excel_path, canonical_day)
        if com_result:
            try:
                return parse_rows_into_signage_data(
                    canonical_day,
                    excel_path,
                    com_result["disp_rows"],
                    com_result["data_rows"],
                    com_result["last_modified"]
                )
            except Exception:
                pass

        # 2. Check file mtime
        try:
            mtime = os.path.getmtime(excel_path)
            with CACHE_LOCK:
                cached = MEMORY_CACHE.get(canonical_day)
                last_mtime = LAST_FILE_MTIME.get(canonical_day)
                if cached and cached.get("success") and last_mtime == mtime:
                    return cached  # File unchanged, reuse memory cache instantly
        except Exception:
            mtime = 0

        # 3. Read file bytes and parse
        file_bytes = read_locked_file_bytes(excel_path)
        if file_bytes is None:
            with CACHE_LOCK:
                return MEMORY_CACHE.get(canonical_day, {
                    "success": False,
                    "day": canonical_day,
                    "error": f"ファイル読込待機中: {os.path.basename(excel_path)}"
                })

        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        sheetnames = wb.sheetnames
        ws_disp = wb[sheetnames[0]]
        ws_data = wb[sheetnames[1]] if len(sheetnames) > 1 else ws_disp
        disp_rows = list(ws_disp.iter_rows(values_only=True))
        data_rows = list(ws_data.iter_rows(values_only=True))
        last_mod = datetime.datetime.fromtimestamp(mtime).strftime("%Y/%m/%d %H:%M:%S")

        parsed_data = parse_rows_into_signage_data(canonical_day, excel_path, disp_rows, data_rows, last_mod)
        with CACHE_LOCK:
            LAST_FILE_MTIME[canonical_day] = mtime
        return parsed_data
    except Exception as e:
        with CACHE_LOCK:
            cached = MEMORY_CACHE.get(canonical_day)
            if cached:
                return cached
        return {"success": False, "day": canonical_day, "error": f"エクセル解析エラー: {str(e)}"}


def render_courses_to_html(courses):
    """Server-side pre-renderer for courses table HTML so it displays fully styled even in previewers."""
    if not courses:
        return '<div style="padding: 40px; text-align: center; color: #64748B;">データがありません</div>'

    def clean_num(val):
        if val is None:
            return ""
        s = str(val).strip()
        if s.endswith(".0"):
            try:
                num = float(s)
                if num.is_integer():
                    s = str(int(num))
            except Exception:
                pass
        return s

    groups = []
    curr_group = None
    for c in courses:
        v = clean_num(c.get("vehicle", ""))
        c["vehicle"] = v
        c["course"] = clean_num(c.get("course", ""))
        if v != "" and curr_group and curr_group["vehicleName"] == v:
            curr_group["courses"].append(c)
        else:
            curr_group = {"vehicleName": v, "courses": [c]}
            groups.append(curr_group)

    now = datetime.datetime.now()
    curr_total_min = now.hour * 60 + now.minute

    def is_within_10min_or_past(t_str):
        if not t_str or ":" not in t_str:
            return False
        try:
            parts = t_str.split(":")
            th = int(parts[0])
            tm = int(parts[1])
            target_min = th * 60 + tm
            return (target_min - curr_total_min) <= 10
        except Exception:
            return False

    html_parts = []
    for group in groups:
        num_courses = len(group["courses"])
        total_rows = num_courses * 2
        is_multi = num_courses > 1

        group_time = ""
        group_comp_time = ""
        group_diff_min = None

        for c in group["courses"]:
            if c.get("time", "").strip():
                group_time = c.get("time", "").strip()
                break

        for c in group["courses"]:
            if c.get("group_completed_time"):
                group_comp_time = c.get("group_completed_time")
            if c.get("group_diff_minutes") is not None:
                group_diff_min = c.get("group_diff_minutes")

        if not group_comp_time and len(group["courses"]) > 0:
            group_comp_time = group["courses"][0].get("course_completed_time", "")
        if group_diff_min is None and len(group["courses"]) > 0:
            group_diff_min = group["courses"][0].get("course_diff_minutes")

        is_group_all_comp = all(c.get("is_completed") is True for c in group["courses"])
        is_group_warn = (not is_group_all_comp) and is_within_10min_or_past(group_time)
        group_warn_class = "time-val-warning" if is_group_warn else ""

        def build_time_cell(time_str, is_comp, comp_time, diff_min, warn_cls):
            diff_html = ""
            if is_comp and (comp_time or diff_min is not None):
                t_label = f"完了 {comp_time}" if comp_time else "完了"
                diff_class = "diff-box-ontime"
                diff_text = "定刻 (±0)"
                if diff_min is not None:
                    if diff_min < 0:
                        diff_class = "diff-box-early"
                        diff_text = f"{diff_min}分 早着"
                    elif diff_min > 0:
                        diff_class = "diff-box-late"
                        diff_text = f"+{diff_min}分 遅延"
                    else:
                        diff_class = "diff-box-ontime"
                        diff_text = "定刻 (±0)"
                diff_html = f'<div class="time-diff-2tier {diff_class}"><span class="diff-tier-time">{t_label}</span><span class="diff-tier-val">{diff_text}</span></div>'
            return f'<div class="time-cell-content"><span class="time-val {warn_cls}">{time_str or ""}</span>{diff_html}</div>'

        html_parts.append('<table class="signage-table vehicle-group-card"><colgroup><col class="col-vehicle"><col class="col-course-sub"><col class="col-time"><col class="col-line"><col class="col-num" span="25"><col class="col-slip"></colgroup><tbody>')

        for idx, c in enumerate(group["courses"]):
            is_first = (idx == 0)
            is_last = (idx == num_courses - 1)
            sep_class = "course-separator-row" if not is_last else ""
            is_nodel = c.get("is_no_delivery", False)

            def build_tile(item):
                t_class = "tile-unstarted"
                d_num = ""
                st = item.get("status", 0)
                if is_nodel:
                    t_class = "tile-green-nodelivery"
                elif st >= 99:
                    t_class = "tile-blue-done"
                    d_num = str(item.get("num", ""))
                elif st == 1:
                    t_class = "tile-grey-active"
                    d_num = str(item.get("num", ""))
                else:
                    t_class = "tile-unstarted"
                return f'<td><div class="cell-tile-container"><div class="progress-tile {t_class}">{d_num}</div></div></td>'

            f_tiles = "".join(build_tile(it) for it in c.get("furidashi", {}).get("items", []))
            s_tiles = "".join(build_tile(it) for it in c.get("sagyo", {}).get("items", []))

            slip_done = c.get("slip", {}).get("is_done", False)
            if is_nodel:
                slip_html = '<div class="slip-checkbox slip-nodelivery-green"></div>'
            elif slip_done:
                slip_html = '<div class="slip-checkbox slip-done-blue">✓</div>'
            else:
                slip_html = '<div class="slip-checkbox slip-pending-empty"></div>'

            grp_comp_cls = "group-completed-cell" if is_group_all_comp else ""
            left_cols = ""
            if is_multi:
                if is_first:
                    t_inner = build_time_cell(group_time, is_group_all_comp, group_comp_time, group_diff_min, group_warn_class)
                    badge_v_cls = "badge-completed" if is_group_all_comp else ""
                    badge_c_cls = "badge-completed" if c.get("is_completed") else ""
                    c_comp_cls = "group-completed-cell" if c.get("is_completed") else ""
                    left_cols = f'''
                    <td class="cell-vehicle-tall {grp_comp_cls}" rowspan="{total_rows}">
                        <div class="badge-vehicle-tall {badge_v_cls}"><span class="badge-text-inner">{group["vehicleName"]}</span></div>
                    </td>
                    <td class="cell-course-sub {c_comp_cls}" rowspan="2">
                        <div class="badge-course-sub {badge_c_cls}"><span class="badge-text-inner">{c.get("course") or "-"}</span></div>
                    </td>
                    <td class="cell-time-tall {grp_comp_cls}" rowspan="{total_rows}" data-time-val="{group_time}" data-is-completed="{str(is_group_all_comp).lower()}">
                        {t_inner}
                    </td>
                    '''
                else:
                    badge_c_cls = "badge-completed" if c.get("is_completed") else ""
                    c_comp_cls = "group-completed-cell" if c.get("is_completed") else ""
                    left_cols = f'''
                    <td class="cell-course-sub {c_comp_cls}" rowspan="2">
                        <div class="badge-course-sub {badge_c_cls}"><span class="badge-text-inner">{c.get("course") or "-"}</span></div>
                    </td>
                    '''
            else:
                is_single_comp = (c.get("is_completed") is True)
                is_single_warn = (not is_single_comp) and is_within_10min_or_past(c.get("time", ""))
                single_warn_cls = "time-val-warning" if is_single_warn else ""
                single_comp_cls = "group-completed-cell" if is_single_comp else ""
                s_comp_time = c.get("group_completed_time") or c.get("course_completed_time") or ""
                s_diff_min = c.get("group_diff_minutes") if c.get("group_diff_minutes") is not None else c.get("course_diff_minutes")
                t_inner = build_time_cell(c.get("time", ""), is_single_comp, s_comp_time, s_diff_min, single_warn_cls)

                def is_hyphen(s):
                    return s is None or str(s).strip() in ("", "-", "ー", "―", "‐", "－", "ｰ")

                is_c_empty = is_hyphen(c.get("course"))
                is_v_empty = is_hyphen(group["vehicleName"])
                is_same = (not is_c_empty) and (c.get("course") == group["vehicleName"])
                if is_c_empty or is_v_empty or is_same:
                    merged_lbl = group["vehicleName"] if not is_v_empty else (c.get("course") if not is_c_empty else "-")
                    badge_full_cls = "badge-completed" if is_single_comp else ""
                    left_cols = f'''
                    <td class="cell-course-full {single_comp_cls}" colspan="2" rowspan="2">
                        <div class="badge-course-full {badge_full_cls}"><span class="badge-text-inner">{merged_lbl}</span></div>
                    </td>
                    '''
                else:
                    badge_v_cls = "badge-completed" if is_single_comp else ""
                    badge_c_cls = "badge-completed" if is_single_comp else ""
                    left_cols = f'''
                    <td class="cell-vehicle-single {single_comp_cls}" rowspan="2">
                        <div class="badge-vehicle-single {badge_v_cls}"><span class="badge-text-inner">{group["vehicleName"] or "-"}</span></div>
                    </td>
                    <td class="cell-course-sub {single_comp_cls}" rowspan="2">
                        <div class="badge-course-sub {badge_c_cls}"><span class="badge-text-inner">{c.get("course") or "-"}</span></div>
                    </td>
                    '''
                left_cols += f'''
                <td class="cell-time {single_comp_cls}" rowspan="2" data-time-val="{c.get("time") or ""}" data-is-completed="{str(is_single_comp).lower()}">
                    {t_inner}
                </td>
                '''

            comp_row_cls = "course-completed-row" if c.get("is_completed") else ""
            f_lbl = c.get("furidashi", {}).get("label", "振出")
            s_lbl = c.get("sagyo", {}).get("label", "査照")

            html_parts.append(f'''
            <tr class="course-row-1 {comp_row_cls}" id="{c.get("id")}_r1">
                {left_cols}
                <td class="cell-line-furidashi">{f_lbl}</td>
                {f_tiles}
                <td class="cell-slip" rowspan="2">
                    <div class="slip-cell-container">{slip_html}</div>
                </td>
            </tr>
            <tr class="course-row-2 {comp_row_cls} {sep_class}" id="{c.get("id")}_r2">
                <td class="cell-line-sagyo">{s_lbl}</td>
                {s_tiles}
            </tr>
            ''')

        html_parts.append('</tbody></table>')

    return "\n".join(html_parts)


def build_standalone_viewer_html(payload):
    """Build a completely self-contained single-file HTML with inline CSS, pre-rendered tables, and embedded scripts."""
    canonical_day = resolve_canonical_day("")
    days_map = payload.get("days", {})
    day_data = days_map.get(canonical_day) or days_map.get("平日") or (list(days_map.values())[0] if days_map else {})
    courses = day_data.get("courses", [])
    count = day_data.get("count", len(courses))
    excel_file = day_data.get("excel_file", "Excel")
    last_mod = day_data.get("last_modified", "")
    time_str = last_mod.split(" ")[1] if " " in last_mod else "--:--:--"

    courses_html = render_courses_to_html(courses)

    css_path = os.path.join(APP_DIR, "style.css")
    css_content = ""
    if os.path.exists(css_path):
        with open(css_path, "r", encoding="utf-8") as f:
            css_content = f.read()

    app_js_path = os.path.join(APP_DIR, "app.js")
    js_content = ""
    if os.path.exists(app_js_path):
        with open(app_js_path, "r", encoding="utf-8") as f:
            js_content = f.read()

    js_data = "window.__ALL_SIGNAGE_DATA__ = " + json.dumps(payload, ensure_ascii=False) + ";\n"
    now_dt = datetime.datetime.now()
    date_jp = now_dt.strftime("%Y/%m/%d")
    days_jp = ["月", "火", "水", "木", "金", "土", "日"]
    weekday_str = days_jp[now_dt.weekday()]
    clock_time_str = now_dt.strftime("%H:%M:%S")

    tab_active_hei = "active" if canonical_day == "平日" else ""
    tab_active_mon = "active" if canonical_day == "月曜" else ""
    tab_active_tue = "active" if canonical_day == "火曜" else ""
    tab_active_sun = "active" if canonical_day == "日・祝" else ""

    html = f"""<!DOCTYPE html>
<html lang="ja">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>倉庫作業進捗サイネージ (閲覧用)</title>
  <link rel="preconnect" href="https://fonts.googleapis.com">
  <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
  <link href="https://fonts.googleapis.com/css2?family=Inter:wght@500;700;800;900&family=Noto+Sans+JP:wght@500;700;900&display=swap" rel="stylesheet">
  <style>
{css_content}
  </style>
</head>
<body>
  <header class="signage-header" id="signage-header">
    <div class="header-left">
      <div class="system-badge">
        <span class="live-dot"></span>
        <span class="badge-text">リアルタイム進捗</span>
      </div>
      <div class="day-tabs-group" id="day-tabs-group">
        <button class="day-tab {tab_active_hei}" data-day="平日">平日</button>
        <button class="day-tab {tab_active_mon}" data-day="月曜">月曜</button>
        <button class="day-tab {tab_active_tue}" data-day="火曜">火曜</button>
        <button class="day-tab {tab_active_sun}" data-day="日・祝">日・祝</button>
      </div>
      <span class="course-count-badge" id="course-count-badge">{count} コース</span>
      <div class="legend-container">
        <div class="legend-item"><div class="legend-tile legend-tile-grey">1</div><span>作業中</span></div>
        <div class="legend-item"><div class="legend-tile legend-tile-blue">1</div><span>完了</span></div>
        <div class="legend-item"><div class="legend-tile legend-tile-green"></div><span>配送なし</span></div>
      </div>
    </div>
    <div class="header-center">
      <div class="clock-display">
        <span class="clock-date" id="clock-date">{date_jp} ({weekday_str})</span>
        <span class="clock-time" id="clock-time">{clock_time_str}</span>
      </div>
    </div>
    <div class="header-right">
      <div class="status-pill" id="sync-status-pill">
        <span class="status-icon" id="sync-status-icon">🟢</span>
        <span class="status-text" id="sync-status-text">同期: {clock_time_str} ({excel_file}: {time_str})</span>
      </div>
      <div class="controls-group">
        <button class="btn-ctrl" id="btn-toggle-scroll" title="スクロール一時停止/再開 (Space)"><span id="scroll-icon">⏸️</span><span id="scroll-btn-text">スクロール中</span></button>
        <button class="btn-ctrl" id="btn-speed" title="スクロール速度切替">⚡ <span id="speed-label">標準</span></button>
        <button class="btn-ctrl" id="btn-settings" title="設定" style="display: none;">⚙️ 設定</button>
        <button class="btn-ctrl btn-fullscreen" id="btn-fullscreen" title="全画面表示 (F11)">⛶ 全画面</button>
      </div>
    </div>
  </header>

  <div class="table-fixed-header-wrapper">
    <table class="signage-table header-table">
      <colgroup>
        <col class="col-vehicle"><col class="col-course-sub"><col class="col-time"><col class="col-line"><col class="col-num" span="25"><col class="col-slip">
      </colgroup>
      <thead>
        <tr>
          <th class="th-course-name th-course-main" id="th-course-main">{canonical_day}コース</th>
          <th class="th-course-name th-course-sub" id="th-course-sub">平日コース</th>
          <th class="th-time">搬送完了時間</th>
          <th class="th-line">項目</th>
          <th class="th-num">1</th><th class="th-num">2</th><th class="th-num">3</th><th class="th-num">4</th><th class="th-num">5</th>
          <th class="th-num">6</th><th class="th-num">7</th><th class="th-num">8</th><th class="th-num">9</th><th class="th-num">10</th>
          <th class="th-num">11</th><th class="th-num">12</th><th class="th-num">13</th><th class="th-num">14</th><th class="th-num">15</th>
          <th class="th-num">16</th><th class="th-num">17</th><th class="th-num">18</th><th class="th-num">19</th><th class="th-num">20</th>
          <th class="th-num">21</th><th class="th-num">22</th><th class="th-num">23</th><th class="th-num">24</th><th class="th-num">25</th>
          <th class="th-slip">伝票</th>
        </tr>
      </thead>
    </table>
  </div>

  <main class="scroll-viewport" id="scroll-viewport">
    <div class="scroll-content" id="scroll-content">
      <div id="course-cards-container">
{courses_html}
      </div>
      <div class="loop-notice-bar" id="loop-notice-bar">
        <span class="loop-spinner"></span>
        <span>全コース表示完了 - まもなく先頭へ戻ります</span>
      </div>
    </div>
  </main>

  <div class="error-banner" id="error-banner" style="display: none;">
    <span class="error-icon">⚠️</span>
    <span class="error-msg" id="error-msg">進捗データを同期中...</span>
  </div>

  <script>
{js_data}
{js_content}
  </script>
</body>
</html>
"""
    return html


LAST_EXPORT_SIGNATURE = ""


def sync_shared_export(force=False):
    """Export current memory cache to signage_data.js and standalone viewer only when data changes."""
    global LAST_EXPORT_SIGNATURE
    try:
        cfg = load_config()
        export_dirs = [APP_DIR]

        # 1. Configured export directories (string or list)
        cfg_export = cfg.get("export_dirs", cfg.get("export_dir", ""))
        if isinstance(cfg_export, list):
            for d in cfg_export:
                if d and os.path.exists(str(d).strip()):
                    p = os.path.abspath(str(d).strip())
                    if p not in [os.path.abspath(x) for x in export_dirs]:
                        export_dirs.append(p)
        elif isinstance(cfg_export, str) and cfg_export.strip():
            p = os.path.abspath(cfg_export.strip())
            if os.path.exists(p) and p not in [os.path.abspath(x) for x in export_dirs]:
                export_dirs.append(p)

        # 2. Teams Excel parent directories
        candidate_excel_paths = [cfg.get("excel_path", "")] + list(cfg.get("excel_paths", {}).values())
        for ep in candidate_excel_paths:
            if ep and os.path.exists(str(ep).strip()):
                parent_d = os.path.abspath(os.path.dirname(str(ep).strip()))
                if os.path.exists(parent_d) and parent_d not in [os.path.abspath(x) for x in export_dirs]:
                    export_dirs.append(parent_d)

        # 3. Auto-detect any ★入力シート folder in OneDrive / Shortcuts
        candidate_onedrive_roots = get_candidate_onedrive_roots()
        for od_root in candidate_onedrive_roots:
            if not os.path.exists(od_root):
                continue
            try:
                for root_dir, dirs, _ in os.walk(od_root):
                    for d in dirs:
                        if "入力シート" in d:
                            target_p = os.path.abspath(os.path.join(root_dir, d))
                            if target_p not in [os.path.abspath(x) for x in export_dirs]:
                                export_dirs.append(target_p)
            except Exception:
                pass

        # Make sure memory cache has all days populated with valid data
        for d in DAYS_ORDER:
            with CACHE_LOCK:
                cached = MEMORY_CACHE.get(d)
            if not cached or not cached.get("success"):
                try:
                    data = refresh_data_for_day(d)
                    with CACHE_LOCK:
                        MEMORY_CACHE[d] = data
                except Exception:
                    pass

        with CACHE_LOCK:
            days_data = {d: MEMORY_CACHE.get(d) for d in DAYS_ORDER if d in MEMORY_CACHE and MEMORY_CACHE.get(d)}

        # MD5 signature of course data to avoid thrashing OneDrive sync
        sig_input = json.dumps(days_data, sort_keys=True, ensure_ascii=False)
        current_sig = hashlib.md5(sig_input.encode("utf-8")).hexdigest()
        if not force and current_sig == LAST_EXPORT_SIGNATURE:
            return  # No change in data! Skip disk writes so OneDrive upload is never interrupted!

        LAST_EXPORT_SIGNATURE = current_sig
        print(f"[EXPORT SYNC] データを書き出しました ({datetime.datetime.now().strftime('%H:%M:%S')})", flush=True)

        payload = {
            "timestamp": datetime.datetime.now().strftime("%Y/%m/%d %H:%M:%S"),
            "config": cfg,
            "days": days_data
        }

        raw_json = json.dumps(payload, ensure_ascii=False)
        js_data = "window.__ALL_SIGNAGE_DATA__ = " + raw_json + ";\n"
        standalone_html = build_standalone_viewer_html(payload)

        for target_dir in export_dirs:
            # 1. Write signage_data.js directly (prevents OneDrive .tmp deletion popups)
            try:
                js_file = os.path.join(target_dir, "signage_data.js")
                with open(js_file, "w", encoding="utf-8") as f:
                    f.write(js_data)
            except Exception:
                pass

            # 2. Write pure JSON (signage_data.json) directly
            try:
                json_file = os.path.join(target_dir, "signage_data.json")
                with open(json_file, "w", encoding="utf-8") as f:
                    f.write(raw_json)
            except Exception:
                pass

            # 3. Write standalone self-contained viewer HTML directly
            try:
                viewer_file = os.path.join(target_dir, "作業進捗サイネージ(閲覧用).html")
                with open(viewer_file, "w", encoding="utf-8") as f:
                    f.write(standalone_html)
            except Exception:
                pass

            # 3. Also copy standard static files if needed
            if os.path.abspath(target_dir) != os.path.abspath(APP_DIR):
                for fname in ["index.html", "style.css", "app.js"]:
                    src = os.path.join(APP_DIR, fname)
                    dst = os.path.join(target_dir, fname)
                    if os.path.exists(src):
                        try:
                            if not os.path.exists(dst) or os.path.getmtime(src) > os.path.getmtime(dst):
                                shutil.copy2(src, dst)
                        except Exception:
                            pass
    except Exception as e:
        print(f"[EXPORT ERROR] {e}", flush=True)


def background_cache_worker():
    """Background polling daemon thread that keeps in-memory data fresh continuously."""
    print("[CACHE ENGINE] 高速インメモリ・キャッシュエンジンが起動しました。", flush=True)

    # Initial load
    for day in DAYS_ORDER:
        try:
            data = refresh_data_for_day(day)
            with CACHE_LOCK:
                MEMORY_CACHE[day] = data
            print(f"[CACHE READY] [{day}] -> {data.get('count', 0)} コース ({data.get('excel_file', '')})", flush=True)
        except Exception as e:
            print(f"[CACHE ERROR] [{day}]: {e}", flush=True)
        time.sleep(0.1)

    # Initial export to shared directories
    sync_shared_export()

    while True:
        try:
            time.sleep(2)
            has_update = False
            for day in DAYS_ORDER:
                try:
                    data = refresh_data_for_day(day)
                    with CACHE_LOCK:
                        MEMORY_CACHE[day] = data
                except Exception:
                    pass
                time.sleep(0.1)
            # Sync to shared storage
            sync_shared_export()
        except Exception:
            time.sleep(2)


class SignageRequestHandler(SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=APP_DIR, **kwargs)

    def log_message(self, format, *args):
        pass

    def end_headers(self):
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Cache-Control", "no-cache, no-store, must-revalidate")
        self.send_header("Pragma", "no-cache")
        self.send_header("Expires", "0")
        super().end_headers()

    def do_GET(self):
        try:
            parsed = urllib.parse.urlparse(self.path)
            if parsed.path == "/api/data":
                query = urllib.parse.parse_qs(parsed.query)
                day_param = query.get("day", [""])[0]
                canonical_day = resolve_canonical_day(day_param)

                # Instant in-memory cache lookup (< 1ms)
                with CACHE_LOCK:
                    data = MEMORY_CACHE.get(canonical_day)

                if not data:
                    data = {
                        "success": True,
                        "day": canonical_day,
                        "title": canonical_day,
                        "count": 0,
                        "courses": [],
                        "last_modified": "初期化中..."
                    }

                cfg = load_config()
                resp_data = dict(data)
                resp_data["config"] = cfg

                resp_json = json.dumps(resp_data, ensure_ascii=False)
                self.send_response(200)
                self.send_header("Content-Type", "application/json; charset=utf-8")
                self.send_header("Access-Control-Allow-Origin", "*")
                self.send_header("Cache-Control", "no-cache, no-store, must-revalidate")
                self.send_header("Pragma", "no-cache")
                self.send_header("Expires", "0")
                self.end_headers()
                self.wfile.write(resp_json.encode("utf-8"))
                return

            if parsed.path == "/api/config":
                cfg = load_config()
                resp_json = json.dumps(cfg, ensure_ascii=False)
                self.send_response(200)
                self.send_header("Content-Type", "application/json; charset=utf-8")
                self.send_header("Access-Control-Allow-Origin", "*")
                self.send_header("Cache-Control", "no-cache, no-store, must-revalidate")
                self.send_header("Pragma", "no-cache")
                self.send_header("Expires", "0")
                self.end_headers()
                self.wfile.write(resp_json.encode("utf-8"))
                return

            super().do_GET()
        except Exception as e:
            try:
                self.send_response(500)
                self.send_header("Content-Type", "application/json; charset=utf-8")
                self.end_headers()
                self.wfile.write(json.dumps({"success": False, "error": str(e)}).encode("utf-8"))
            except Exception:
                pass

    def do_POST(self):
        try:
            parsed = urllib.parse.urlparse(self.path)
            if parsed.path == "/api/config":
                content_length = int(self.headers.get("Content-Length", 0))
                body = self.rfile.read(content_length)
                try:
                    new_cfg = json.loads(body.decode("utf-8"))
                    current_cfg = load_config()
                    updated_cfg = {**current_cfg, **new_cfg}
                    save_config(updated_cfg)

                    self.send_response(200)
                    self.send_header("Content-Type", "application/json; charset=utf-8")
                    self.end_headers()
                    self.wfile.write(json.dumps({"success": True, "config": updated_cfg}).encode("utf-8"))
                except Exception as e:
                    self.send_response(400)
                    self.send_header("Content-Type", "application/json; charset=utf-8")
                    self.end_headers()
                    self.wfile.write(json.dumps({"success": False, "error": str(e)}).encode("utf-8"))
                return

            self.send_response(404)
            self.end_headers()
        except Exception:
            pass


def main():
    cfg = load_config()
    port = int(cfg.get("port", PORT))
    print("=" * 60, flush=True)
    print("  倉庫作業進捗サイネージシステム (高速インメモリ・複数曜日対応)", flush=True)
    print(f"  起動URL: http://localhost:{port}", flush=True)
    print("  ※この画面を閉じるとサイネージが停止します", flush=True)
    print("=" * 60, flush=True)

    try:
        ThreadingHTTPServer.allow_reuse_address = True
        server = ThreadingHTTPServer(("", port), SignageRequestHandler)
        server.daemon_threads = True
        print(f"[SERVER] サーバーがポート {port} で正常に起動しました (即時受付可能)。", flush=True)
    except Exception as e:
        print(f"[SERVER FATAL ERROR] {e}", flush=True)
        return

    # Start background polling cache thread
    cache_thread = threading.Thread(target=background_cache_worker, daemon=True)
    cache_thread.start()

    while True:
        try:
            server.serve_forever()
        except Exception as e:
            print(f"[SERVER EXCEPTION] {e}", flush=True)
            time.sleep(1)
        except KeyboardInterrupt:
            print("\nサーバーを停止しました。", flush=True)
            break


if __name__ == "__main__":
    main()
